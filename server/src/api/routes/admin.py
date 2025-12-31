from datetime import date
from typing import Optional, List
from fastapi import APIRouter, BackgroundTasks, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
import csv
import io
from sqlalchemy.orm import Session

# Excel support (optional - gracefully degrade if not installed)
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORTED = True
except ImportError:
    EXCEL_SUPPORTED = False

from src.api.deps import AuthContext, get_auth_context, get_db
from src.services.admin_service import AdminService
from src.schemas.admin import (
    OrganizationCreate,
    OrganizationUpdate,
    OrganizationOut,
    OrderDetailOut,
    MeOut,
    LabelsIn,
    LabelOut,
    CsvImportOut,
    DashboardOut,
    OrderUpdate,
    NotificationListOut,
    NotificationStats,
    BulkTokenRequest,
    BulkTokenResponse,
    AnalyticsOut,
    BulkTokenResult,
    ReminderRequest,
    ReminderResponse,
)
from src.schemas.order import OrderCreate, OrderOut
from src.models.organization import Organization

router = APIRouter(prefix="/admin", tags=["admin"])


def _is_org_admin(role: Optional[str]) -> bool:
    if not role:
        return False
    r = role.strip().lower()
    if r in {"admin", "owner"}:
        return True
    # Clerk often uses "org:admin" / "org:basic_member"
    if r.endswith(":admin") or r.endswith(":owner"):
        return True
    return False


@router.get("/me", response_model=MeOut)
def get_me(
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    org = None
    if ctx.organization_id is not None:
        org = db.query(Organization).filter(Organization.id == ctx.organization_id).first()
    return MeOut(
        sub=ctx.sub,
        org_external_id=ctx.org_external_id,
        org_role=ctx.org_role,
        organization=org,
    )


@router.get("/org", response_model=OrganizationOut)
def get_org_settings(
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    org = db.query(Organization).filter(Organization.id == ctx.organization_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="ORG_NOT_FOUND")
    return org


@router.put("/org", response_model=OrganizationOut)
def update_org_settings(
    payload: OrganizationUpdate,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    # Admin-key always allowed; otherwise only org admins.
    if not (ctx.is_admin_key or _is_org_admin(ctx.org_role)):
        raise HTTPException(status_code=403, detail="FORBIDDEN")
    return AdminService(db).update_organization(ctx.organization_id, payload)


@router.get("/organizations", response_model=list[OrganizationOut])
def list_organizations(
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    # Bearer token users get their own org only (multi-tenant).
    # Admin-key without X-Org-Id can see all orgs (dev/platform).
    return AdminService(db).list_organizations(scope_org_id=ctx.organization_id)


@router.post("/organizations", response_model=OrganizationOut)
def create_organization(
    payload: OrganizationCreate,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    if not ctx.is_admin_key:
        # In production, org is provisioned via external IdP (e.g. Clerk Organizations).
        # Keep this endpoint for admin-key only.
        raise HTTPException(status_code=403, detail="FORBIDDEN")
    return AdminService(db).create_organization(payload)


@router.get("/orders")
def list_orders(
    organization_id: Optional[int] = Query(default=None),
    q: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    day: Optional[str] = Query(default=None, description="YYYY-MM-DD (Asia/Seoul)"),
    today: bool = Query(default=False),
    start_date: Optional[date] = Query(default=None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(default=None, description="End date (YYYY-MM-DD)"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=100),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    # Tenant scoping: bearer token forces ctx.organization_id.
    org_id = ctx.organization_id if ctx.organization_id is not None else organization_id
    return AdminService(db).list_orders(
        organization_id=org_id,
        q=q,
        status=status,
        day=day,
        today=today,
        start_date=start_date,
        end_date=end_date,
        page=page,
        limit=limit,
    )


@router.post("/orders/import/csv", response_model=CsvImportOut)
def import_orders_csv(
    file: UploadFile = File(...),
    strict: bool = Query(default=False),
    auto_generate_tokens: bool = Query(default=True, description="Automatically generate QR tokens for imported orders"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Import orders from a CSV file.

    Expected header columns (case-insensitive):
      - order_number (or order_no)
      - context (optional)
      - sender_name (or buyer_name)
      - sender_phone (or buyer_phone)
      - recipient_name (or receiver_name) (optional)
      - recipient_phone (or receiver_phone) (optional)

    When auto_generate_tokens=True (default), QR tokens are automatically
    generated for all successfully imported orders.
    """

    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")

    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="CSV_FILE_REQUIRED")

    raw = file.file.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        # allow cp949 in KR field ops
        text = raw.decode("cp949", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for r in reader:
        # normalize keys to snake_lower
        nr = {str(k).strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in (r or {}).items()}
        rows.append(nr)

    svc = AdminService(db)
    created_ids, errors = svc.import_orders_csv(rows=rows, organization_id=ctx.organization_id, strict=strict)

    # Auto-generate tokens for imported orders
    generated_tokens_count = 0
    if auto_generate_tokens and created_ids:
        token_result = svc.bulk_generate_tokens(
            order_ids=created_ids,
            scope_org_id=ctx.organization_id,
            force=False,
        )
        generated_tokens_count = token_result.get("success_count", 0)

    return {
        "created_count": len(created_ids),
        "created_order_ids": created_ids,
        "errors": errors,
        "generated_tokens_count": generated_tokens_count,
    }


def _parse_excel_to_rows(raw: bytes) -> list[dict]:
    """Parse Excel file to list of normalized row dicts."""
    if not EXCEL_SUPPORTED:
        raise HTTPException(status_code=400, detail="EXCEL_NOT_SUPPORTED: openpyxl not installed")

    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        val = cell.value
        if val:
            headers.append(str(val).strip().lower())
        else:
            headers.append("")

    # Parse data rows
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                if val is not None:
                    row_dict[headers[i]] = str(val).strip() if isinstance(val, str) else str(val)
                else:
                    row_dict[headers[i]] = ""
        if row_dict:
            rows.append(row_dict)

    return rows


@router.post("/orders/import", response_model=CsvImportOut)
def import_orders(
    file: UploadFile = File(...),
    strict: bool = Query(default=False),
    auto_generate_tokens: bool = Query(default=True, description="Automatically generate QR tokens"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Import orders from CSV or Excel file (auto-detected by extension).

    Supported formats:
      - CSV (.csv): UTF-8 or CP949 encoding
      - Excel (.xlsx): First sheet is used

    Expected columns (case-insensitive):
      - order_number (or order_no) - Required
      - context (optional)
      - sender_name (or buyer_name) - Required
      - sender_phone (or buyer_phone) - Required
      - recipient_name (or receiver_name) (optional)
      - recipient_phone (or receiver_phone) (optional)
    """
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")

    if not file.filename:
        raise HTTPException(status_code=400, detail="FILE_REQUIRED")

    filename_lower = file.filename.lower()
    raw = file.file.read()

    # Parse based on file extension
    if filename_lower.endswith(".xlsx"):
        rows = _parse_excel_to_rows(raw)
    elif filename_lower.endswith(".csv"):
        try:
            text = raw.decode("utf-8-sig")
        except Exception:
            text = raw.decode("cp949", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for r in reader:
            nr = {str(k).strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in (r or {}).items()}
            rows.append(nr)
    else:
        raise HTTPException(status_code=400, detail="UNSUPPORTED_FORMAT: Use .csv or .xlsx files")

    svc = AdminService(db)
    created_ids, errors = svc.import_orders_csv(rows=rows, organization_id=ctx.organization_id, strict=strict)

    generated_tokens_count = 0
    if auto_generate_tokens and created_ids:
        token_result = svc.bulk_generate_tokens(
            order_ids=created_ids,
            scope_org_id=ctx.organization_id,
            force=False,
        )
        generated_tokens_count = token_result.get("success_count", 0)

    return {
        "created_count": len(created_ids),
        "created_order_ids": created_ids,
        "errors": errors,
        "generated_tokens_count": generated_tokens_count,
    }


@router.post("/orders", response_model=OrderOut)
def create_order(
    payload: OrderCreate,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).create_order(payload, organization_id=ctx.organization_id)


@router.get("/orders/{order_id}", response_model=OrderDetailOut)
def get_order_detail(
    order_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    return AdminService(db).get_order_detail(order_id, scope_org_id=ctx.organization_id)


@router.post("/orders/{order_id}/token")
def issue_order_token(
    order_id: int,
    force: bool = Query(default=False),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    return AdminService(db).issue_token(order_id, scope_org_id=ctx.organization_id, force=force)


@router.post("/orders/{order_id}/notify")
async def resend_notification(
    order_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    return await AdminService(db).resend_notification(order_id, background_tasks, scope_org_id=ctx.organization_id)


@router.post("/orders/labels", response_model=list[LabelOut])
def get_labels(
    payload: LabelsIn,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).get_labels(
        order_ids=payload.order_ids,
        scope_org_id=ctx.organization_id,
        ensure_tokens=payload.ensure_tokens,
        force=payload.force,
    )


@router.patch("/orders/{order_id}", response_model=OrderOut)
def update_order(
    order_id: int,
    payload: OrderUpdate,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Update an existing order."""
    return AdminService(db).update_order(order_id, payload, scope_org_id=ctx.organization_id)


@router.delete("/orders/{order_id}")
def delete_order(
    order_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Delete an order and all related data."""
    return AdminService(db).delete_order(order_id, scope_org_id=ctx.organization_id)


@router.get("/dashboard", response_model=DashboardOut)
def get_dashboard(
    start_date: Optional[date] = Query(default=None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Get dashboard KPI and recent proofs."""
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).get_dashboard(
        organization_id=ctx.organization_id,
        start_date=start_date,
        end_date=end_date,
    )


@router.get("/notifications", response_model=NotificationListOut)
def list_notifications(
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=100),
    status: Optional[str] = Query(default=None, description="SENT, FAILED, PENDING"),
    channel: Optional[str] = Query(default=None, description="ALIMTALK, SMS"),
    start_date: Optional[date] = Query(default=None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """List notifications with pagination and filters."""
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).list_notifications(
        organization_id=ctx.organization_id,
        page=page,
        limit=limit,
        status=status,
        channel=channel,
        start_date=start_date,
        end_date=end_date,
    )


@router.get("/notifications/stats", response_model=NotificationStats)
def get_notification_stats(
    start_date: Optional[date] = Query(default=None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Get notification statistics (success/failed/pending counts)."""
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).get_notification_stats(
        organization_id=ctx.organization_id,
        start_date=start_date,
        end_date=end_date,
    )


@router.post("/orders/bulk-tokens", response_model=BulkTokenResponse)
def bulk_generate_tokens(
    payload: BulkTokenRequest,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Generate tokens for multiple orders at once.

    WARNING: If force=True, existing tokens will be replaced (breaks previously shared links).
    """
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).bulk_generate_tokens(
        order_ids=payload.order_ids,
        scope_org_id=ctx.organization_id,
        force=payload.force,
    )


@router.get("/orders/export/csv")
def export_orders_csv(
    status: Optional[str] = Query(default=None),
    start_date: Optional[date] = Query(default=None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Export orders to CSV file.

    Returns a streaming CSV response with order data.
    """
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")

    csv_data = AdminService(db).export_orders_csv(
        organization_id=ctx.organization_id,
        status=status,
        start_date=start_date,
        end_date=end_date,
    )

    return StreamingResponse(
        iter([csv_data]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=orders_{date.today().isoformat()}.csv"
        },
    )


@router.get("/analytics", response_model=AnalyticsOut)
def get_analytics(
    start_date: Optional[date] = Query(default=None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Get detailed analytics with trends and breakdowns.

    If no dates provided, defaults to last 30 days.
    """
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).get_analytics(
        organization_id=ctx.organization_id,
        start_date=start_date,
        end_date=end_date,
    )


@router.post("/orders/reminders", response_model=ReminderResponse)
async def send_reminders(
    payload: ReminderRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Send reminder notifications to orders pending proof upload.

    - order_ids: Optional list of specific order IDs. If None, finds all pending orders.
    - hours_since_token: Only remind orders where token was issued > N hours ago.
    - max_reminders: Maximum number of reminders per order (prevents spam).
    """
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return await AdminService(db).send_reminders(
        organization_id=ctx.organization_id,
        order_ids=payload.order_ids,
        hours_since_token=payload.hours_since_token,
        max_reminders=payload.max_reminders,
        background_tasks=background_tasks,
    )


@router.get("/orders/pending-reminders")
def get_pending_reminders(
    hours_since_token: int = Query(default=24, description="Token issued > N hours ago"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_auth_context),
):
    """Get list of orders that could receive reminders.

    Returns orders that:
    - Have token issued but no proof uploaded
    - Token was issued > hours_since_token ago
    """
    if ctx.organization_id is None:
        raise HTTPException(status_code=403, detail="ORG_REQUIRED")
    return AdminService(db).get_pending_reminders(
        organization_id=ctx.organization_id,
        hours_since_token=hours_since_token,
    )
